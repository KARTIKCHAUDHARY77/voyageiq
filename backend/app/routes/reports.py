"""
VoyageIQ AI — Maritime Intelligence Platform
Copyright (c) 2024 Kartik Chaudhary. All Rights Reserved.
Unauthorized copying or use of this file is strictly prohibited.
Contact: 2512520007@geu.ac.in
"""

"""
VoyageIQ AI - Report Generation Blueprint
Generates PDF (ReportLab), Excel (openpyxl), and CSV voyage reports.
"""
import io
import csv
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file, current_app
from flask_jwt_extended import jwt_required, get_jwt_identity

from app.extensions import db
from app.models import Voyage, Vessel, NoonReport, Claim, FuelAnalytic

reports_bp = Blueprint('reports', __name__)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_float(val, default=0.0) -> float:
    try:
        return float(val) if val is not None else default
    except (TypeError, ValueError):
        return default


def _fmt(val, decimals=1, unit=''):
    if val is None:
        return 'N/A'
    try:
        formatted = f"{float(val):.{decimals}f}"
        return f"{formatted} {unit}".strip() if unit else formatted
    except (TypeError, ValueError):
        return 'N/A'


def _load_voyage_data(voyage_id: str):
    """Load all data needed to build a voyage report."""
    voyage = Voyage.query.get(voyage_id)
    if not voyage:
        return None, None, [], [], []
    vessel = Vessel.query.get(voyage.vessel_id)
    reports = (
        NoonReport.query
        .filter_by(voyage_id=voyage_id)
        .order_by(NoonReport.report_date.asc())
        .all()
    )
    claims = Claim.query.filter_by(voyage_id=voyage_id).all()
    fuel_analytics = FuelAnalytic.query.filter_by(voyage_id=voyage_id).all()
    return voyage, vessel, reports, claims, fuel_analytics


def _compute_kpis(voyage, reports, claims):
    """Compute executive summary KPIs."""
    total_fuel = sum(_safe_float(r.total_fuel_consumption) for r in reports)
    avg_speed = (
        sum(_safe_float(r.speed_over_ground) for r in reports) / len(reports)
        if reports else 0
    )
    total_distance = sum(_safe_float(r.distance_noon_to_noon) for r in reports)
    avg_wind_bft = (
        sum(r.wind_force_bft or 0 for r in reports) / len(reports)
        if reports else 0
    )
    total_claim_impact = sum(_safe_float(c.estimated_impact_usd) for c in claims)
    open_claims = [c for c in claims if c.status == 'open']
    fuel_efficiency = total_distance / total_fuel if total_fuel > 0 else 0

    cp_speed = _safe_float(voyage.charter_party_speed)
    cp_consumption = _safe_float(voyage.charter_party_consumption)
    speed_variance = round(avg_speed - cp_speed, 2) if cp_speed else None

    return {
        'total_fuel_mt': round(total_fuel, 2),
        'avg_speed_kn': round(avg_speed, 2),
        'total_distance_nm': round(total_distance, 1),
        'avg_wind_bft': round(avg_wind_bft, 1),
        'total_claim_impact_usd': round(total_claim_impact, 2),
        'open_claims_count': len(open_claims),
        'fuel_efficiency_nm_per_mt': round(fuel_efficiency, 3),
        'report_count': len(reports),
        'cp_speed': cp_speed,
        'cp_consumption': cp_consumption,
        'speed_variance': speed_variance,
        'performance_score': _safe_float(voyage.performance_score),
    }


# ---------------------------------------------------------------------------
# PDF Report (ReportLab)
# ---------------------------------------------------------------------------

def _generate_pdf(voyage, vessel, reports, claims, fuel_analytics) -> bytes:
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            HRFlowable, PageBreak,
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    except ImportError:
        raise ImportError("reportlab is not installed. Run: pip install reportlab")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2.5 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    NAVY = colors.HexColor('#1e3a5f')
    TEAL = colors.HexColor('#0d9488')
    LIGHT_GREY = colors.HexColor('#f1f5f9')
    MID_GREY = colors.HexColor('#94a3b8')

    style_title = ParagraphStyle('Title', parent=styles['Title'],
                                 fontSize=22, textColor=NAVY, spaceAfter=6)
    style_h1 = ParagraphStyle('H1', parent=styles['Heading1'],
                               fontSize=14, textColor=NAVY, spaceBefore=14, spaceAfter=4)
    style_h2 = ParagraphStyle('H2', parent=styles['Heading2'],
                               fontSize=11, textColor=TEAL, spaceBefore=10, spaceAfter=3)
    style_body = ParagraphStyle('Body', parent=styles['Normal'], fontSize=9, spaceAfter=3)
    style_small = ParagraphStyle('Small', parent=styles['Normal'], fontSize=8,
                                 textColor=MID_GREY)

    kpis = _compute_kpis(voyage, reports, claims)
    vessel_name = vessel.name if vessel else 'Unknown Vessel'
    voyage_num = voyage.voyage_number
    dep_port = voyage.departure_port
    arr_port = voyage.arrival_port
    report_date = datetime.utcnow().strftime('%d %B %Y')

    elements = []

    # ---- Cover Page ------------------------------------------------------ #
    elements.append(Spacer(1, 1.5 * cm))
    elements.append(Paragraph("VoyageIQ Maritime Intelligence", style_small))
    elements.append(Paragraph("VOYAGE PERFORMANCE REPORT", style_title))
    elements.append(HRFlowable(width='100%', thickness=2, color=NAVY))
    elements.append(Spacer(1, 0.4 * cm))

    cover_data = [
        ["Vessel Name", vessel_name, "Report Date", report_date],
        ["IMO Number", vessel.imo_number if vessel else 'N/A', "Voyage Number", voyage_num],
        ["Vessel Type", vessel.vessel_type if vessel else 'N/A', "Status", voyage.status.replace('_', ' ').title()],
        ["Departure Port", dep_port, "Arrival Port", arr_port],
        ["ETD", voyage.etd.strftime('%d %b %Y') if voyage.etd else 'N/A',
         "ETA", voyage.eta.strftime('%d %b %Y') if voyage.eta else 'N/A'],
        ["Charterer", voyage.charterer or 'N/A', "CP Speed", f"{_fmt(voyage.charter_party_speed, 1)} kn"],
    ]
    cover_tbl = Table(cover_data, colWidths=[4 * cm, 6.5 * cm, 4 * cm, 5 * cm])
    cover_tbl.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (-1, -1), LIGHT_GREY),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, LIGHT_GREY]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(cover_tbl)
    elements.append(Spacer(1, 0.8 * cm))

    # ---- Executive Summary KPIs ----------------------------------------- #
    elements.append(Paragraph("Executive Summary", style_h1))
    elements.append(HRFlowable(width='100%', thickness=1, color=TEAL))
    elements.append(Spacer(1, 0.3 * cm))

    kpi_data = [
        ["Metric", "Value", "Metric", "Value"],
        ["Total Fuel Consumed", f"{kpis['total_fuel_mt']:.1f} MT",
         "Avg Speed (SOG)", f"{kpis['avg_speed_kn']:.2f} kn"],
        ["Total Distance", f"{kpis['total_distance_nm']:.1f} nm",
         "Fuel Efficiency", f"{kpis['fuel_efficiency_nm_per_mt']:.3f} nm/MT"],
        ["Avg Wind (Beaufort)", f"Bft {kpis['avg_wind_bft']:.1f}",
         "Performance Score", f"{kpis['performance_score']:.1f}/100"],
        ["Open Claims", str(kpis['open_claims_count']),
         "Total Claim Impact", f"${kpis['total_claim_impact_usd']:,.0f}"],
        ["Noon Reports", str(kpis['report_count']),
         "Speed vs CP", f"{kpis['speed_variance']:+.2f} kn" if kpis['speed_variance'] is not None else 'N/A'],
    ]
    kpi_tbl = Table(kpi_data, colWidths=[5.5 * cm, 4.5 * cm, 5.5 * cm, 4.5 * cm])
    kpi_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), NAVY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_GREY]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(kpi_tbl)
    elements.append(PageBreak())

    # ---- Daily Performance Table ---------------------------------------- #
    elements.append(Paragraph("Daily Performance Analysis", style_h1))
    elements.append(HRFlowable(width='100%', thickness=1, color=TEAL))
    elements.append(Spacer(1, 0.3 * cm))

    perf_headers = ["Date", "SOG (kn)", "Dist (nm)", "RPM", "Fuel (MT)", "Wind Bft", "Wave (m)", "ROB LSFO"]
    perf_data = [perf_headers]
    for r in reports:
        perf_data.append([
            r.report_date.strftime('%d %b %Y') if r.report_date else 'N/A',
            _fmt(r.speed_over_ground, 1),
            _fmt(r.distance_noon_to_noon, 1),
            _fmt(r.rpm, 0),
            _fmt(r.total_fuel_consumption, 2),
            str(r.wind_force_bft or 'N/A'),
            _fmt(r.wave_height, 1),
            _fmt(r.rob_lsfo, 1),
        ])

    perf_tbl = Table(perf_data, colWidths=[2.5 * cm, 1.8 * cm, 1.8 * cm, 1.6 * cm,
                                            1.8 * cm, 1.8 * cm, 1.8 * cm, 2.3 * cm])
    perf_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), NAVY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_GREY]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(perf_tbl)
    elements.append(PageBreak())

    # ---- Fuel Analysis -------------------------------------------------- #
    elements.append(Paragraph("Fuel Analysis", style_h1))
    elements.append(HRFlowable(width='100%', thickness=1, color=TEAL))
    elements.append(Spacer(1, 0.3 * cm))

    fuel_headers = ["Date", "ME LSFO (MT)", "ME MGO (MT)", "AE MGO (MT)", "Boiler (MT)", "Total (MT)"]
    fuel_data = [fuel_headers]
    for r in reports:
        total = _safe_float(r.total_fuel_consumption) or (
            _safe_float(r.me_lsfo) + _safe_float(r.me_mgo) +
            _safe_float(r.ae_mgo) + _safe_float(r.boiler_lsfo) + _safe_float(r.boiler_mgo)
        )
        fuel_data.append([
            r.report_date.strftime('%d %b %Y') if r.report_date else 'N/A',
            _fmt(r.me_lsfo, 2),
            _fmt(r.me_mgo, 2),
            _fmt(r.ae_mgo, 2),
            _fmt(r.boiler_lsfo, 2),
            f"{total:.2f}",
        ])

    fuel_tbl = Table(fuel_data, colWidths=[2.8 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm])
    fuel_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), TEAL),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_GREY]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(fuel_tbl)
    elements.append(Spacer(1, 0.5 * cm))

    # ---- Claims Summary ------------------------------------------------- #
    if claims:
        elements.append(Paragraph("Claims Summary", style_h1))
        elements.append(HRFlowable(width='100%', thickness=1, color=TEAL))
        elements.append(Spacer(1, 0.3 * cm))
        claim_headers = ["Type", "Severity", "Status", "Period Start", "Period End", "Impact (USD)"]
        claim_data = [claim_headers]
        for c in claims:
            claim_data.append([
                c.claim_type or 'N/A',
                (c.severity or 'N/A').upper(),
                (c.status or 'N/A').replace('_', ' ').title(),
                c.period_start.strftime('%d %b %Y') if c.period_start else 'N/A',
                c.period_end.strftime('%d %b %Y') if c.period_end else 'N/A',
                f"${_safe_float(c.estimated_impact_usd):,.0f}",
            ])
        claim_tbl = Table(claim_data, colWidths=[3.5 * cm, 2 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 2.8 * cm])
        claim_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fef2f2')]),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#fca5a5')),
            ('ALIGN', (5, 0), (5, -1), 'RIGHT'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(claim_tbl)
        elements.append(Spacer(1, 0.5 * cm))

    # ---- Recommendations ------------------------------------------------ #
    elements.append(Paragraph("Recommendations", style_h1))
    elements.append(HRFlowable(width='100%', thickness=1, color=TEAL))
    elements.append(Spacer(1, 0.3 * cm))

    sv = kpis.get('speed_variance')
    recs = []
    if sv is not None and sv < -0.5:
        recs.append(f"Speed is {abs(sv):.2f} kn below CP warranted speed. Investigate hull condition and engine performance.")
    if kpis['avg_wind_bft'] > 5:
        recs.append("Average Beaufort > 5 indicates sustained adverse weather. Verify charter party weather adjustments are applied.")
    if kpis['open_claims_count'] > 0:
        recs.append(f"{kpis['open_claims_count']} open claim(s) with total exposure ${kpis['total_claim_impact_usd']:,.0f}. Compile supporting documentation.")
    recs += [
        "Schedule hull underwater inspection if more than 6 months since last cleaning.",
        "Review trim optimisation to target maximum fuel efficiency.",
        "Implement JIT arrival protocols to eliminate anchor waiting at full speed.",
    ]

    for i, rec in enumerate(recs, 1):
        elements.append(Paragraph(f"{i}. {rec}", style_body))

    elements.append(Spacer(1, 0.5 * cm))
    elements.append(HRFlowable(width='100%', thickness=0.5, color=MID_GREY))
    elements.append(Paragraph(
        f"Report generated by VoyageIQ AI Maritime Intelligence Platform — {report_date} UTC",
        style_small
    ))

    doc.build(elements)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Excel Report (openpyxl)
# ---------------------------------------------------------------------------

def _generate_excel(voyage, vessel, reports, claims, fuel_analytics) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is not installed. Run: pip install openpyxl")

    wb = Workbook()

    NAVY_HEX = "1E3A5F"
    TEAL_HEX = "0D9488"
    LIGHT_HEX = "F1F5F9"
    RED_HEX = "DC2626"

    def _hdr_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _hdr_font(color="FFFFFF", bold=True, size=10):
        return Font(name='Calibri', bold=bold, color=color, size=size)

    def _body_font(size=9):
        return Font(name='Calibri', size=size)

    def _thin_border():
        s = Side(style='thin', color='CBD5E1')
        return Border(left=s, right=s, top=s, bottom=s)

    def _write_header_row(ws, headers, row=1, fill_hex=NAVY_HEX):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = _hdr_fill(fill_hex)
            cell.font = _hdr_font()
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = _thin_border()

    def _write_data_row(ws, values, row, alt=False):
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = _body_font()
            cell.border = _thin_border()
            if alt:
                cell.fill = PatternFill("solid", fgColor=LIGHT_HEX)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _auto_width(ws, min_w=10, max_w=30):
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=min_w)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, min_w), max_w)

    kpis = _compute_kpis(voyage, reports, claims)
    vessel_name = vessel.name if vessel else 'Unknown Vessel'

    # ---- Sheet 1: Summary ---------------------------------------------- #
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.row_dimensions[1].height = 30

    ws_sum['A1'] = "VoyageIQ — Voyage Performance Report"
    ws_sum['A1'].font = Font(name='Calibri', bold=True, size=16, color=NAVY_HEX)
    ws_sum.merge_cells('A1:D1')

    summary_rows = [
        ("Vessel", vessel_name), ("IMO Number", vessel.imo_number if vessel else 'N/A'),
        ("Voyage Number", voyage.voyage_number), ("Status", voyage.status),
        ("Departure", voyage.departure_port), ("Arrival", voyage.arrival_port),
        ("ETD", voyage.etd.strftime('%d %b %Y') if voyage.etd else 'N/A'),
        ("ETA", voyage.eta.strftime('%d %b %Y') if voyage.eta else 'N/A'),
        ("Charterer", voyage.charterer or 'N/A'),
        ("CP Speed", f"{_fmt(voyage.charter_party_speed, 1)} kn"),
        ("CP Consumption", f"{_fmt(voyage.charter_party_consumption, 1)} MT/day"),
        ("", ""),
        ("Total Fuel (MT)", kpis['total_fuel_mt']),
        ("Avg Speed (kn)", kpis['avg_speed_kn']),
        ("Total Distance (nm)", kpis['total_distance_nm']),
        ("Fuel Efficiency (nm/MT)", kpis['fuel_efficiency_nm_per_mt']),
        ("Performance Score", kpis['performance_score']),
        ("Open Claims", kpis['open_claims_count']),
        ("Total Claim Impact ($)", kpis['total_claim_impact_usd']),
        ("Noon Reports", kpis['report_count']),
    ]

    for i, (label, value) in enumerate(summary_rows, start=3):
        ws_sum.cell(row=i, column=1, value=label).font = Font(bold=True, size=9)
        ws_sum.cell(row=i, column=2, value=value).font = Font(size=9)

    _auto_width(ws_sum)

    # ---- Sheet 2: Daily Performance ------------------------------------- #
    ws_perf = wb.create_sheet("Daily Performance")
    perf_headers = ["Date", "SOG (kn)", "STW (kn)", "Distance (nm)", "DTG (nm)",
                     "RPM", "ME Power (kW)", "Wind (Bft)", "Wave (m)", "Swell (m)", "Course"]
    _write_header_row(ws_perf, perf_headers)
    for i, r in enumerate(reports, start=2):
        _write_data_row(ws_perf, [
            r.report_date.isoformat() if r.report_date else None,
            _safe_float(r.speed_over_ground) or None,
            _safe_float(r.speed_through_water) or None,
            _safe_float(r.distance_noon_to_noon) or None,
            _safe_float(r.distance_to_go) or None,
            _safe_float(r.rpm) or None,
            _safe_float(r.me_power) or None,
            r.wind_force_bft,
            _safe_float(r.wave_height) or None,
            _safe_float(r.swell_height) or None,
            _safe_float(r.course) or None,
        ], row=i, alt=(i % 2 == 0))
    _auto_width(ws_perf)

    # ---- Sheet 3: Fuel Data --------------------------------------------- #
    ws_fuel = wb.create_sheet("Fuel Data")
    fuel_headers = ["Date", "ME LSFO (MT)", "ME MGO (MT)", "AE LSFO (MT)", "AE MGO (MT)",
                     "Boiler LSFO (MT)", "Boiler MGO (MT)", "Total Fuel (MT)", "ROB LSFO (MT)", "ROB MGO (MT)"]
    _write_header_row(ws_fuel, fuel_headers, fill_hex=TEAL_HEX)
    for i, r in enumerate(reports, start=2):
        total = _safe_float(r.total_fuel_consumption) or (
            _safe_float(r.me_lsfo) + _safe_float(r.me_mgo) +
            _safe_float(r.ae_lsfo) + _safe_float(r.ae_mgo) +
            _safe_float(r.boiler_lsfo) + _safe_float(r.boiler_mgo)
        )
        _write_data_row(ws_fuel, [
            r.report_date.isoformat() if r.report_date else None,
            _safe_float(r.me_lsfo) or None,
            _safe_float(r.me_mgo) or None,
            _safe_float(r.ae_lsfo) or None,
            _safe_float(r.ae_mgo) or None,
            _safe_float(r.boiler_lsfo) or None,
            _safe_float(r.boiler_mgo) or None,
            round(total, 3) if total else None,
            _safe_float(r.rob_lsfo) or None,
            _safe_float(r.rob_mgo) or None,
        ], row=i, alt=(i % 2 == 0))
    _auto_width(ws_fuel)

    # ---- Sheet 4: Claims ------------------------------------------------ #
    ws_claims = wb.create_sheet("Claims")
    claim_headers = ["Claim Type", "Severity", "Status", "Detected Date",
                      "Period Start", "Period End", "Expected", "Actual",
                      "Variance", "Unit", "Impact (USD)", "Description"]
    _write_header_row(ws_claims, claim_headers, fill_hex=RED_HEX)
    for i, c in enumerate(claims, start=2):
        _write_data_row(ws_claims, [
            c.claim_type,
            c.severity,
            c.status,
            c.detected_date.isoformat() if c.detected_date else None,
            c.period_start.isoformat() if c.period_start else None,
            c.period_end.isoformat() if c.period_end else None,
            _safe_float(c.expected_value) or None,
            _safe_float(c.actual_value) or None,
            _safe_float(c.variance) or None,
            c.unit,
            _safe_float(c.estimated_impact_usd) or None,
            c.description or '',
        ], row=i, alt=(i % 2 == 0))
    _auto_width(ws_claims)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# CSV Report
# ---------------------------------------------------------------------------

def _generate_csv(voyage, vessel, reports, claims, fuel_analytics) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)

    # Header metadata
    writer.writerow(["VoyageIQ Voyage Report"])
    writer.writerow(["Vessel", vessel.name if vessel else 'N/A',
                      "Voyage", voyage.voyage_number,
                      "Route", f"{voyage.departure_port} → {voyage.arrival_port}"])
    writer.writerow(["Generated", datetime.utcnow().isoformat()])
    writer.writerow([])

    # Noon report rows
    writer.writerow([
        "report_date", "latitude", "longitude", "speed_over_ground", "speed_through_water",
        "distance_noon_to_noon", "distance_to_go", "rpm", "course", "wind_force_bft",
        "wind_direction", "wind_speed_knots", "wave_height", "swell_height",
        "me_lsfo", "me_mgo", "ae_lsfo", "ae_mgo", "boiler_lsfo", "boiler_mgo",
        "total_fuel_consumption", "rob_lsfo", "rob_mgo", "cargo_quantity",
        "draft_fore", "draft_aft", "me_power", "rpm",
    ])

    for r in reports:
        writer.writerow([
            r.report_date.isoformat() if r.report_date else '',
            _safe_float(r.latitude), _safe_float(r.longitude),
            _safe_float(r.speed_over_ground), _safe_float(r.speed_through_water),
            _safe_float(r.distance_noon_to_noon), _safe_float(r.distance_to_go),
            _safe_float(r.rpm), _safe_float(r.course), r.wind_force_bft or '',
            r.wind_direction or '', _safe_float(r.wind_speed_knots),
            _safe_float(r.wave_height), _safe_float(r.swell_height),
            _safe_float(r.me_lsfo), _safe_float(r.me_mgo),
            _safe_float(r.ae_lsfo), _safe_float(r.ae_mgo),
            _safe_float(r.boiler_lsfo), _safe_float(r.boiler_mgo),
            _safe_float(r.total_fuel_consumption),
            _safe_float(r.rob_lsfo), _safe_float(r.rob_mgo),
            _safe_float(r.cargo_quantity),
            _safe_float(r.draft_fore), _safe_float(r.draft_aft),
            _safe_float(r.me_power), _safe_float(r.rpm),
        ])

    return buf.getvalue().encode('utf-8')


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@reports_bp.route('/generate/<voyage_id>', methods=['GET'])
@jwt_required()
def generate_report(voyage_id):
    """
    GET /api/reports/generate/<voyage_id>?format=pdf|excel|csv
    Generate and download a voyage report in the requested format.
    """
    try:
        fmt = request.args.get('format', 'pdf').lower()
        if fmt not in ('pdf', 'excel', 'csv'):
            return jsonify({'success': False, 'error': "format must be one of: pdf, excel, csv"}), 400

        voyage, vessel, reports, claims, fuel_analytics = _load_voyage_data(voyage_id)
        if not voyage:
            return jsonify({'success': False, 'error': 'Voyage not found.'}), 404

        vessel_name_safe = (vessel.name if vessel else 'vessel').replace(' ', '_')
        voyage_num_safe = voyage.voyage_number.replace(' ', '_')
        base_filename = f"VoyageIQ_{vessel_name_safe}_{voyage_num_safe}"

        if fmt == 'pdf':
            try:
                data = _generate_pdf(voyage, vessel, reports, claims, fuel_analytics)
            except ImportError as exc:
                return jsonify({'success': False, 'error': str(exc)}), 500
            return send_file(
                io.BytesIO(data),
                mimetype='application/pdf',
                as_attachment=True,
                download_name=f"{base_filename}.pdf",
            )

        elif fmt == 'excel':
            try:
                data = _generate_excel(voyage, vessel, reports, claims, fuel_analytics)
            except ImportError as exc:
                return jsonify({'success': False, 'error': str(exc)}), 500
            return send_file(
                io.BytesIO(data),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f"{base_filename}.xlsx",
            )

        else:  # csv
            data = _generate_csv(voyage, vessel, reports, claims, fuel_analytics)
            return send_file(
                io.BytesIO(data),
                mimetype='text/csv',
                as_attachment=True,
                download_name=f"{base_filename}.csv",
            )

    except Exception as exc:
        current_app.logger.error(f'Report generation error: {exc}', exc_info=True)
        return jsonify({'success': False, 'error': 'Report generation failed.'}), 500


@reports_bp.route('/templates', methods=['GET'])
@jwt_required()
def list_templates():
    """GET /api/reports/templates — Return available report templates."""
    templates = [
        {
            "id": "voyage_summary",
            "name": "Voyage Performance Report",
            "description": "Full voyage report with cover page, KPI summary, daily performance, fuel analysis, and claims.",
            "formats": ["pdf", "excel", "csv"],
            "sections": ["Cover Page", "Executive Summary KPIs", "Daily Performance Analysis",
                         "Fuel Analysis", "Claims Summary", "Recommendations"],
        },
        {
            "id": "fuel_only",
            "name": "Fuel Consumption Report",
            "description": "Detailed fuel consumption breakdown across all fuel types and ROB.",
            "formats": ["excel", "csv"],
            "sections": ["Fuel Consumption by Day", "ROB Tracking", "Efficiency Metrics"],
        },
        {
            "id": "claims",
            "name": "Claims Report",
            "description": "Full claims register with severity, status, and financial impact.",
            "formats": ["pdf", "excel"],
            "sections": ["Open Claims", "Closed Claims", "Total Exposure Summary"],
        },
        {
            "id": "performance",
            "name": "Performance Report",
            "description": "Speed, RPM, and weather-adjusted performance versus charter party.",
            "formats": ["pdf", "excel", "csv"],
            "sections": ["Speed Analysis", "CP Variance", "Weather Factors", "Recommendations"],
        },
    ]
    return jsonify({'success': True, 'templates': templates, 'count': len(templates)}), 200
